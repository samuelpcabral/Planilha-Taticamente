from openpyxl import load_workbook
from pandas import read_html


class CreateExcel:

    def __init__(self):
        pass

    @staticmethod
    def read_attributes_html(html_file):
        is_gk = False
        player_attributes = dict()
        tables = read_html(html_file, encoding='UTF-8')
        if tables[0].columns[0] == "Guarda-Redes":
            is_gk = True
        values = tables[0].columns[-1]

        if is_gk:
            table_guarda_redes = tables[0]
            dict_guarda_redes = {
                'Saídas da Baliza': int(table_guarda_redes.loc[table_guarda_redes["Guarda-Redes"] == "(Tendência) para Saídas da Baliza", values]),
                'Para Socar': int(table_guarda_redes.loc[table_guarda_redes["Guarda-Redes"] == "(Tendência) para Socar", values]),
                'Alcance Aéreo': int(table_guarda_redes.loc[table_guarda_redes["Guarda-Redes"] == "Alcance Aéreo", values]),
                'Comando de Área': int(table_guarda_redes.loc[table_guarda_redes["Guarda-Redes"] == "Comando de Área", values]),
                'Comunicação': int(table_guarda_redes.loc[table_guarda_redes["Guarda-Redes"] == "Comunicação", values]),
                'Excentricidade': int(table_guarda_redes.loc[table_guarda_redes["Guarda-Redes"] == "Excentricidade", values]),
                'Jogo de Mãos': int(table_guarda_redes.loc[table_guarda_redes["Guarda-Redes"] == "Jogo de Mãos", values]),
                'Lançamentos': int(table_guarda_redes.loc[table_guarda_redes["Guarda-Redes"] == "Lançamentos", values]),
                'Passe': int(table_guarda_redes.loc[table_guarda_redes["Guarda-Redes"] == "Passe", values]),
                'Pontapé': int(table_guarda_redes.loc[table_guarda_redes["Guarda-Redes"] == "Pontapé", values]),
                'Primeiro Toque': int(table_guarda_redes.loc[table_guarda_redes["Guarda-Redes"] == "Primeiro Toque", values]),
                'Reflexos': int(table_guarda_redes.loc[table_guarda_redes["Guarda-Redes"] == "Reflexos", values]),
                'Um Para Um': int(table_guarda_redes.loc[table_guarda_redes["Guarda-Redes"] == "Um Para Um", values]),
            }
            player_attributes['Guarda-Redes'] = dict_guarda_redes
        else:
            table_tecnicos = tables[0]
            dict_tecnicos = {
                'Cabeceamento': int(table_tecnicos.loc[table_tecnicos["Técnicos"] == "Cabeceamento", values]),
                'Cantos': int(table_tecnicos.loc[table_tecnicos["Técnicos"] == "Cantos", values]),
                'Cruzamentos': int(table_tecnicos.loc[table_tecnicos["Técnicos"] == "Cruzamentos", values]),
                'Desarme': int(table_tecnicos.loc[table_tecnicos["Técnicos"] == "Desarme", values]),
                'Finalização': int(table_tecnicos.loc[table_tecnicos["Técnicos"] == "Finalização", values]),
                'Finta': int(table_tecnicos.loc[table_tecnicos["Técnicos"] == "Finta", values]),
                'Lançamentos Longos': int(table_tecnicos.loc[table_tecnicos["Técnicos"] == "Lançamentos Longos", values]),
                'Livres': int(table_tecnicos.loc[table_tecnicos["Técnicos"] == "Livres", values]),
                'Marcação': int(table_tecnicos.loc[table_tecnicos["Técnicos"] == "Marcação", values]),
                'Marcação de Penáltis': int(table_tecnicos.loc[table_tecnicos["Técnicos"] == "Marcação de Penáltis", values]),
                'Passe': int(table_tecnicos.loc[table_tecnicos["Técnicos"] == "Passe", values]),
                'Primeiro Toque': int(table_tecnicos.loc[table_tecnicos["Técnicos"] == "Primeiro Toque", values]),
                'Remates de Longe': int(table_tecnicos.loc[table_tecnicos["Técnicos"] == "Remates de Longe", values]),
                'Técnica': int(table_tecnicos.loc[table_tecnicos["Técnicos"] == "Técnica", values])
            }
            player_attributes['Tecnicos'] = dict_tecnicos
        table_mentais = tables[1]
        dict_mentais = {
            'Agressividade': int(table_mentais.loc[table_mentais["Mentais"] == "Agressividade", values]),
            'Antecipação': int(table_mentais.loc[table_mentais["Mentais"] == "Antecipação", values]),
            'Bravura': int(table_mentais.loc[table_mentais["Mentais"] == "Bravura", values]),
            'Compostura': int(table_mentais.loc[table_mentais["Mentais"] == "Compostura", values]),
            'Concentração': int(table_mentais.loc[table_mentais["Mentais"] == "Concentração", values]),
            'Decisões': int(table_mentais.loc[table_mentais["Mentais"] == "Decisões", values]),
            'Determinação': int(table_mentais.loc[table_mentais["Mentais"] == "Determinação", values]),
            'Imprevisibilidade': int(table_mentais.loc[table_mentais["Mentais"] == "Imprevisibilidade", values]),
            'Índice de Trabalho': int(table_mentais.loc[table_mentais["Mentais"] == "Índice de Trabalho", values]),
            'Liderança': int(table_mentais.loc[table_mentais["Mentais"] == "Liderança", values]),
            'Posicionamento': int(table_mentais.loc[table_mentais["Mentais"] == "Posicionamento", values]),
            'Sem Bola': int(table_mentais.loc[table_mentais["Mentais"] == "Sem Bola", values]),
            'Trabalho de Equipa': int(table_mentais.loc[table_mentais["Mentais"] == "Trabalho de Equipa", values]),
            'Visão de Jogo': int(table_mentais.loc[table_mentais["Mentais"] == "Visão de Jogo", values])
        }
        player_attributes['Mentais'] = dict_mentais
        table_fisicos = tables[2]
        dict_fisicos = {
            "Aceleração": int(table_fisicos.loc[table_fisicos["Físicos"] == "Aceleração", values]),
            "Agilidade": int(table_fisicos.loc[table_fisicos["Físicos"] == "Agilidade", values]),
            "Aptidão Física": int(table_fisicos.loc[table_fisicos["Físicos"] == "Aptidão Física", values]),
            "Equilíbrio": int(table_fisicos.loc[table_fisicos["Físicos"] == "Equilíbrio", values]),
            "Força": int(table_fisicos.loc[table_fisicos["Físicos"] == "Força", values]),
            "Impulsão": int(table_fisicos.loc[table_fisicos["Físicos"] == "Impulsão", values]),
            "Resistência": int(table_fisicos.loc[table_fisicos["Físicos"] == "Resistência", values]),
            "Velocidade": int(table_fisicos.loc[table_fisicos["Físicos"] == "Velocidade", values])
        }
        player_attributes['Fisicos'] = dict_fisicos
        return player_attributes, is_gk

    def fill_excel(self, template, new_name, attributes, is_gk):
        wb = load_workbook(template)
        atributos = wb['Atributos']
        wb.active = atributos

        if is_gk:
            atributos['B42'] = attributes['Guarda-Redes']['Saídas da Baliza']
            atributos['B43'] = attributes['Guarda-Redes']['Para Socar']
            atributos['B44'] = attributes['Guarda-Redes']['Alcance Aéreo']
            atributos['B45'] = attributes['Guarda-Redes']['Comando de Área']
            atributos['B46'] = attributes['Guarda-Redes']['Comunicação']
            atributos['B47'] = attributes['Guarda-Redes']['Excentricidade']
            atributos['B48'] = attributes['Guarda-Redes']['Jogo de Mãos']
            atributos['B49'] = attributes['Guarda-Redes']['Lançamentos']
            atributos['B50'] = attributes['Guarda-Redes']['Passe']
            atributos['B51'] = attributes['Guarda-Redes']['Pontapé']
            atributos['B52'] = attributes['Guarda-Redes']['Primeiro Toque']
            atributos['B53'] = attributes['Guarda-Redes']['Reflexos']
            atributos['B54'] = attributes['Guarda-Redes']['Um Para Um']
        else:
            atributos['B3'] = attributes['Tecnicos']['Cabeceamento']
            atributos['B4'] = attributes['Tecnicos']['Cantos']
            atributos['B5'] = attributes['Tecnicos']['Cruzamentos']
            atributos['B6'] = attributes['Tecnicos']['Desarme']
            atributos['B7'] = attributes['Tecnicos']['Finalização']
            atributos['B8'] = attributes['Tecnicos']['Finta']
            atributos['B9'] = attributes['Tecnicos']['Lançamentos Longos']
            atributos['B10'] = attributes['Tecnicos']['Livres']
            atributos['B11'] = attributes['Tecnicos']['Marcação']
            atributos['B12'] = attributes['Tecnicos']['Marcação de Penáltis']
            atributos['B13'] = attributes['Tecnicos']['Passe']
            atributos['B14'] = attributes['Tecnicos']['Primeiro Toque']
            atributos['B15'] = attributes['Tecnicos']['Remates de Longe']
            atributos['B16'] = attributes['Tecnicos']['Técnica']

        atributos['B18'] = attributes['Mentais']['Agressividade']
        atributos['B19'] = attributes['Mentais']['Antecipação']
        atributos['B20'] = attributes['Mentais']['Bravura']
        atributos['B21'] = attributes['Mentais']['Compostura']
        atributos['B22'] = attributes['Mentais']['Concentração']
        atributos['B23'] = attributes['Mentais']['Decisões']
        atributos['B24'] = attributes['Mentais']['Determinação']
        atributos['B25'] = attributes['Mentais']['Imprevisibilidade']
        atributos['B26'] = attributes['Mentais']['Índice de Trabalho']
        atributos['B27'] = attributes['Mentais']['Liderança']
        atributos['B28'] = attributes['Mentais']['Posicionamento']
        atributos['B29'] = attributes['Mentais']['Sem Bola']
        atributos['B30'] = attributes['Mentais']['Trabalho de Equipa']
        atributos['B31'] = attributes['Mentais']['Visão de Jogo']

        atributos['B33'] = attributes['Fisicos']['Aceleração']
        atributos['B34'] = attributes['Fisicos']['Agilidade']
        atributos['B35'] = attributes['Fisicos']['Aptidão Física']
        atributos['B36'] = attributes['Fisicos']['Equilíbrio']
        atributos['B37'] = attributes['Fisicos']['Força']
        atributos['B38'] = attributes['Fisicos']['Impulsão']
        atributos['B39'] = attributes['Fisicos']['Resistência']
        atributos['B40'] = attributes['Fisicos']['Velocidade']

        wb.save(f'{new_name}.xlsx')


def create_sheets(template, files):
    ce = CreateExcel()
    for file in files:
        attributes, is_gk = ce.read_attributes_html(file)
        filename = file.rsplit('/')
        new_filename = filename[-1].replace('.html', '')
        ce.fill_excel(template, new_filename, attributes, is_gk)
    return str(len(files))
